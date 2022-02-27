import os
from os import listdir, path
from openpyxl import load_workbook
from docxtpl import DocxTemplate
from num2words import num2words
import datetime
#import win32com.client

from tkinter import *
from tkinter.filedialog import askopenfilenames
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
import tkinter.font as TkFont
from PIL import ImageTk, Image
from tkinter import Tk, Canvas, Entry, Text, Button, PhotoImage


def num_two_words(a):
    # change numbers to words in desired format
    a = str(a)
    a = a.replace(" ", "")
    num = (a.split('.'))
    num_word = int(num[0])
    if num_word > 0:
        num_word = num2words(num_word, lang='pl')
        decimal = num[1]
        if decimal == '0':
            return f'{num_word} i {decimal}0/100 PLN'
        else:
           return f'{num_word} i {decimal}/100 PLN'


def number_name(n):
    # funcion for file name prorpouse
    n = n.replace('-', '_')
    n = n.replace('/', '_')
    n = n[5:-3]
    return n


class CreateDocument:

    def __init__(self, master, filename_xlsx, filename_doc, user):
        self.master = master
        self.master.title('CreateDocument')
        self.filename_xlsx = filename_xlsx
        self.filename_doc = filename_doc
        self.user = user
        self.table_col_names= self.load_data_base()
        self.table_col_names.append('Użytkownik')





        self.label_1 = Label(master,
                           text='Upewnij się, że użyłeś zmiennych w szablonie z poniższej listy:',
                           bg='#2c2f33',
                           fg='#FFFFFF',
                           font=('Rubik', 20),
                      )

        self.label_1.place(
            x=200.0,
            y=100.0,

        )

        self.label_2 = Label(master,
                           text=self.table_col_names,
                           bg='#2c2f33',
                           fg='#FFFFFF',
                           font=('Rubik', 18),
                           highlightbackground = "#25CEDE",
                           highlightcolor= "#25CEDE",
                           highlightthickness=4,
                           borderwidth=3,
                           )

        self.label_2.config(text=("\n".join(self.table_col_names)))

        self.label_2.place(
            x=350.0,
            y=180.0,

        )

        self.button_1 = Button(master,
                             text ='Jest OK, wykonaj',
                             bg='#433E3E',
                             fg='#FFFFFF',
                             font=('Rubik', 25),
                             highlightbackground ="#25CEDE",
                             highlightcolor="#25CEDE",
                             highlightthickness=4,
                             borderwidth=3,
                             command =self.main,
                             relief="raised",
                    )

        self.button_1.place(
            x=200.0,
            y=800.0,
            width=500.0,
            height=100.0
        )

        self.button_2 = Button(master,
                             text ='Zamknij',
                             bg='#433E3E',
                             fg='#FFFFFF',
                             font=('Rubik', 25),
                             highlightbackground ="#25CEDE",
                             highlightcolor="#25CEDE",
                             highlightthickness=4,
                             borderwidth=3,
                             command=master.quit,
                             relief="raised",
                    )

        self.button_2.place(
            x=800.0,
            y=800.0,
            width=200.0,
            height=100.0
        )

    def value_row_index(self):
        # find row index that is filled in
        column_list1 = []
        for row in self.sheet.iter_rows(min_row=1, max_col=self.sheet.max_column):
            for cell in row:
              if(cell.value is not None):
                column_list1.append(cell.value)
                if cell.value == column_list1[0]:
                   self.val_row_index = cell.row
                   return self.val_row_index


    def get_row_val(self, row_number):
        # to get column titles
        self.column_list2 = []
        for cell in self.sheet[row_number]:
            if(cell.value is not None):
                self.column_list2.append(cell.value)
        return self.column_list2

    def load_data_base(self):
        # to load dates from known base
        self.workbook = load_workbook(self.filename_xlsx)
        self.sheet = self.workbook.active
        number = self.value_row_index()
        self.column_list2 = []
        for cell in self.sheet[number]:
            if(cell.value is not None):
                self.column_list2.append(cell.value)
        self.table_col_names = self.column_list2
        self.table_col_names = [i.replace(' ', '_') for i in self.table_col_names]
        return self.table_col_names





    def get_dataformat(self):
        # to replace datatime in self.content dict
        self.data_dict = {k: v for k, v in self.content.items() if type(v) is datetime.datetime }
        for k, v in self.data_dict.items():
            # do something with value
            new_value = v.strftime('%m.%d.%Y')
            self.data_dict[k] = new_value
        return self.data_dict



    def financial_data(self):
        # to select data to financial operations
        self.content['Wydatki_niekwalifikowalne'] = self.content['Wartość_ogółem'] - self.content['Wydatki_kwalifikowalne']
        self.financial_dict = {k: "{:,.2f}".format(v).replace(',', ' ') for k, v in self.content.items() if isinstance(v, (int, float, complex))}
        print(self.financial_dict)
        return(self.financial_dict)

    def word_num(self):
        # to change numbers to words and add new dict with wodr_num_values
        self.financial_data()
        self.d = {k+'_słownie':v for k,v in self.financial_dict.items()}
        for k, v in self.d.items():
            # do something with value
            word_value = num_two_words(v)

            self.d[k] = word_value

            return self.d

    def create_content(self, values_list):
        # to create a main content for .docx needs
        # values_list is a list of values from each row from database
        keys_list = self.load_data_base()
        self.content = dict(zip(keys_list, values_list))
        print('zip dict', self.content)
        self.main_content = self.content.copy()
        self.get_dataformat()
        self.main_content.update(self.data_dict) # x.update(d) update dict x o wartości dict d
        key = 'Wartość_ogółem'
        if key in self.main_content:
            self.financial_data()
            self.main_content.update(self.financial_dict)
            self.word_num()
            self.main_content.update(self.d)
        return self.main_content

    def create_word_doc(self):
        # to create word.docx document from ready template
        doc = DocxTemplate(self.filename_doc)
        self.context = self.main_content
        self.context['Użytkownik'] = self.user
        doc.render(self.context)
        self.column_names = self.context.keys()
        key = 'Tytuł_pliku'
        if key in self.context:
            print('halooo')
            filename = "{}.docx".format(self.context['Tytuł_pliku'])
            print('name', filename)
        else:
            a = self.context['Numer_projektu']
            b = self.context['Beneficjent']
            filename = "{}_{}.docx".format(number_name(a), b)
        doc.save(filename)


    def main(self):
        # create and save documents
        self.load_data_base()
        number = self.value_row_index()
        for i in range(number + 1, self.sheet.max_row + 1):
            values_list = self.get_row_val(i)
            self.create_content(values_list)
            self.create_word_doc()



    def mail_body(self):
        self.outlook = win32com.client.Dispatch('outlook.application')
        self.mail = outlook.CreateItem(0)
        self.load_data_base()
        number = self.value_row_index()
        for i in range(number + 1, self.sheet.max_row + 1):
            values_list = self.get_row_val(i)
            self.create_content(values_list)
            self.context = self.main_content
            self.mail.To = self.context['Email']
            self.mail.Subject = self.context['Tytuł']
            self.mail.HTMLBody = '<h3>To będzie treść maila</h3>'
            self.mail.Body = "This is the normal body"
            self.mail.Attachments.Add(self.context['zal_1'])
            self.mail.Attachments.Add(self.context['zal_2'])
            self.mail.Display()


class StartWindow:

    def __init__(self, master, users):
        self.master = master
        self.master.title('DPR')
        self.users = users


# logo and title image
        self.image1 = Image.open("image_1.png")
        self.test = ImageTk.PhotoImage(self.image1)
        self.label1 = Label(bg = "#2C2F33",image=self.test)
        self.label1.image = self.test
        self.label1.place(x=200.0, y=120.0)

        self.image2 = Image.open("entry_1.png")
        self.test = ImageTk.PhotoImage(self.image2)
        self.label2 = Label(bg = "#2C2F33",image=self.test)
        self.label2.image = self.test
        self.label2.place(x=730.0, y=130.0)

# menu to choose an welcome_user

        def show(event):
            # function to get and welcom user
            self.user = self.clicked.get()
            self.welcome_user = self.clicked.get().split(' ', 1)[0]
            text = 'Witaj {}, co dzisiaj robimy ?'.format(self.welcome_user)
            welcome_label = Label(self.master,
                                  text=text,
                                  font=('Rubik', 25),
                                  bg='#2C2F33', fg='#FFFFFF'
                               )
            welcome_label.place(x=130.0, y=580.0)


        self.clicked = StringVar()
        self.clicked.set('Wybierz użytkownika')

        self.drop = OptionMenu(self.master,
            self.clicked,
            *self.users,
            command=show
        )

        self.drop.configure(bg='#433E3E',
            fg='#FFFFFF',
            font=('Rubik', 25),
            highlightbackground="#25CEDE",
            highlightcolor="#25CEDE",
            highlightthickness=4,
            borderwidth=3,
            relief="raised",
        )

        self.drop.place(
            x=619.0,
            y=326.0,
            width=557.0,
            height=77.0
        )

        self.menu = self.master.nametowidget(self.drop.menuname)
        self.menu.config(font=('Rubik', 18 ), bg='#433E3E', fg='#FFFFFF')


# create buttons to start choosen program
        button_1 = Button(
            self.master,
            text='Stwórz dokument',
            bg='#433E3E',
            fg='#FFFFFF',
            font=('Rubik', 25),
            highlightbackground = "#25CEDE",
            highlightcolor= "#25CEDE",
            highlightthickness=4,
            borderwidth=3,
            command=self.run_1,
            relief="raised"
       )
        button_1.place(
            x=130.0,
            y=689.0,
            width=357.0,
            height=147.0
        )

        button_2 = Button(
            self.master,
            text='Stwórz emaile, under costruction...',
            bg='#433E3E',
            fg='#FFFFFF',
            font=('Rubik', 25),
            highlightbackground="#25CEDE",
            highlightcolor= "#25CEDE",
            highlightthickness=4,
            borderwidth=3,
            command=self.run_2,
            relief="raised"
       )
        button_2.place(
            x=600.0,
            y=689.0,
            width=357.0,
            height=147.0
        )


    def run_1(self):

        filename_xlsx = filedialog.askopenfilename(title='Wybierz plik Twojej baza_danych w .xlsx', filetypes=[('XLSX files', '*.xlsx')])
        filename_doc = filedialog.askopenfilename(title='Wybierz szablon pisma .docx', filetypes=[('DOCX files', '*.docx')])
        user = self.user
        self.newWindow = Toplevel(self.master, bg="#2C2F33", height=1000, width=1200)
        self.app = CreateDocument(self.newWindow, filename_xlsx, filename_doc, user)
        self.app.load_data_base()


    def run_2(self):

        filename_xlsx = filedialog.askopenfilename(title='Wybierz plik Twojej baza_danych w .xlsx', filetypes=[('XLSX files', '*.xlsx')])
        filename_doc = 'None'
        self.newWindow = Toplevel(self.master, bg="#2C2F33", height=1000, width=1200)
        self.app = CreateDocument(self.newWindow, filename_xlsx, filename_doc, user)
        self.app.mail_body()





# main program #
if __name__ == "__main__":

    users = ['MR',
             'ML',
             'ES',
            
         ]

    root = Tk()
    root.geometry("1440x1024")
    root.configure(bg = "#2C2F33")
    app = StartWindow(root, users)
    root.mainloop()
